"""
AUDIT TRAIL EXPORT (GeoX)

Export audit trails to PDF and Excel for SAMREC/JORC compliance.
"""

from __future__ import annotations

from pathlib import Path
from datetime import datetime
from typing import Optional

import pandas as pd

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.pdfgen import canvas
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

from .drillhole_audit_trail import AuditTrail

# Export availability flags for UI components
__all__ = ['export_to_excel', 'export_to_pdf', 'export_to_csv', 'OPENPYXL_AVAILABLE', 'REPORTLAB_AVAILABLE']


def export_to_excel(audit_trail: AuditTrail, file_path: str) -> bool:
    """
    Export audit trail to Excel with formatting.
    
    Args:
        audit_trail: AuditTrail instance
        file_path: Output file path
        
    Returns:
        True if successful, False otherwise
    """
    if not OPENPYXL_AVAILABLE:
        raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
    
    df = audit_trail.to_dataframe()
    if df.empty:
        # Create empty workbook with headers
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Audit Trail"
        ws.append(["No changes recorded in this session."])
        wb.save(file_path)
        return True
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit Trail"
    
    # Header style
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Data styles
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Write summary sheet
    summary_ws = wb.create_sheet("Summary", 0)
    summary = audit_trail.get_summary()
    
    summary_ws.append(["AUDIT TRAIL SUMMARY"])
    summary_ws.append([])
    summary_ws.append(["Project:", audit_trail.project_name])
    summary_ws.append(["Session Start:", audit_trail.session_start.strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append(["Session End:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
    summary_ws.append(["Session Duration:", summary["session_duration"]])
    summary_ws.append([])
    summary_ws.append(["Total Changes:", summary["total_changes"]])
    summary_ws.append(["Auto-Fixes:", summary["auto_fixes"]])
    summary_ws.append(["Manual Edits:", summary["manual_edits"]])
    summary_ws.append(["Batch Edits:", summary["batch_edits"]])
    summary_ws.append(["Find & Replace:", summary["find_replace"]])
    summary_ws.append([])
    summary_ws.append(["Tables Affected:", ", ".join(sorted(summary["tables_affected"]))])
    summary_ws.append(["Holes Affected:", len(summary["holes_affected"])])
    
    # Format summary sheet
    summary_ws.column_dimensions['A'].width = 20
    summary_ws.column_dimensions['B'].width = 40
    for row in range(1, summary_ws.max_row + 1):
        cell = summary_ws[f'A{row}']
        if row == 1:
            cell.font = Font(bold=True, size=14)
        elif row <= 6:
            cell.font = Font(bold=True)
    
    # Write audit trail data
    # Headers
    headers = list(df.columns)
    ws.append(headers)
    
    # Style headers
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data
    for row_idx, row_data in enumerate(df.itertuples(index=False), start=2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = border
            cell.alignment = left_align if col_idx <= 6 else center_align
            
            # Highlight auto-fixes
            if col_idx == 2 and value == "Auto-Fix":  # Change Type column
                cell.fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    
    # Auto-adjust column widths
    for col_idx, col in enumerate(df.columns, 1):
        max_length = max(
            len(str(col)),
            df[col].astype(str).map(len).max() if not df.empty else 0
        )
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Save
    wb.save(file_path)
    return True


def export_to_pdf(audit_trail: AuditTrail, file_path: str) -> bool:
    """
    Export audit trail to PDF with professional formatting.
    
    Args:
        audit_trail: AuditTrail instance
        file_path: Output file path
        
    Returns:
        True if successful, False otherwise
    """
    if not REPORTLAB_AVAILABLE:
        raise ImportError("reportlab is required for PDF export. Install with: pip install reportlab")
    
    df = audit_trail.to_dataframe()
    summary = audit_trail.get_summary()
    
    # Create PDF document
    doc = SimpleDocTemplate(
        file_path,
        pagesize=letter,
        rightMargin=0.5*inch,
        leftMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch,
    )
    
    # Container for PDF elements
    story = []
    styles = getSampleStyleSheet()
    
    # Title
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        textColor=colors.HexColor('#1a1a1a'),
        spaceAfter=30,
        alignment=1,  # Center
    )
    story.append(Paragraph("DRILLHOLE DATA AUDIT TRAIL", title_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Summary section
    summary_style = ParagraphStyle(
        'Summary',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=12,
    )
    
    story.append(Paragraph("<b>Project:</b> " + audit_trail.project_name, summary_style))
    story.append(Paragraph("<b>Session Start:</b> " + audit_trail.session_start.strftime("%Y-%m-%d %H:%M:%S"), summary_style))
    story.append(Paragraph("<b>Session End:</b> " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"), summary_style))
    story.append(Paragraph("<b>Session Duration:</b> " + summary["session_duration"], summary_style))
    story.append(Spacer(1, 0.1*inch))
    
    story.append(Paragraph("<b>Summary Statistics:</b>", summary_style))
    story.append(Paragraph(f"Total Changes: {summary['total_changes']}", summary_style))
    story.append(Paragraph(f"Auto-Fixes: {summary['auto_fixes']}", summary_style))
    story.append(Paragraph(f"Manual Edits: {summary['manual_edits']}", summary_style))
    story.append(Paragraph(f"Batch Edits: {summary['batch_edits']}", summary_style))
    story.append(Paragraph(f"Find & Replace: {summary['find_replace']}", summary_style))
    story.append(Paragraph(f"Tables Affected: {', '.join(sorted(summary['tables_affected']))}", summary_style))
    story.append(Paragraph(f"Holes Affected: {len(summary['holes_affected'])}", summary_style))
    story.append(Spacer(1, 0.2*inch))
    
    # Audit trail table
    if df.empty:
        story.append(Paragraph("<i>No changes recorded in this session.</i>", summary_style))
    else:
        story.append(Paragraph("<b>Detailed Audit Trail:</b>", summary_style))
        story.append(Spacer(1, 0.1*inch))
        
        # Prepare table data
        table_data = [list(df.columns)]  # Headers
        
        # Add data rows (limit to prevent huge PDFs)
        max_rows = 1000
        for idx, row in df.iterrows():
            if idx >= max_rows:
                table_data.append([f"... ({len(df) - max_rows} more rows not shown)"])
                break
            table_data.append([str(val) for val in row])
        
        # Create table
        table = Table(table_data, repeatRows=1)
        table.setStyle(TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 9),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f5f5')]),
        ]))
        
        story.append(table)
    
    # Footer with page numbers
    def add_page_number(canvas_obj, doc):
        canvas_obj.saveState()
        canvas_obj.setFont('Helvetica', 9)
        page_num = canvas_obj.getPageNumber()
        text = f"Page {page_num}"
        canvas_obj.drawCentredString(4.25*inch, 0.5*inch, text)
        canvas_obj.restoreState()
    
    # Build PDF
    doc.build(story, onFirstPage=add_page_number, onLaterPages=add_page_number)
    return True


def export_to_csv(audit_trail: AuditTrail, file_path: str) -> bool:
    """
    Export audit trail to CSV.
    
    Args:
        audit_trail: AuditTrail instance
        file_path: Output file path
        
    Returns:
        True if successful, False otherwise
    """
    df = audit_trail.to_dataframe()
    if df.empty:
        # Create empty CSV with headers
        df = pd.DataFrame(columns=[
            "Timestamp", "Change Type", "User", "Table", "Hole ID", "Row Index",
            "Column", "Old Value", "New Value", "Rule Code", "Reason", "Confidence"
        ])
    
    df.to_csv(file_path, index=False)
    return True

